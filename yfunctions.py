import pandas as pd
import numpy as np
import traceback
from sqlalchemy import create_engine
from tabulate import tabulate
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
import inspect
import os
import sys
import re
import unicodedata
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import datetime
from dateutil.relativedelta import relativedelta
from urllib.request import urlretrieve
import zipfile
import shutil

################ SQL FUNCTIONS ################


def login(server, database, username, password):
    conn = 'mssql+pyodbc://'+username+':'+password+'@'+server + \
        '/'+database+'?driver=SQL+Server+Native+Client+11.0'
    engine = create_engine(conn, echo=False, pool_size=50, max_overflow=100)
    return [conn, engine]


def search_table(substring, conn):

    sql_tables = pd.read_sql('SELECT * FROM information_schema.tables', conn)
    sql_tables = sql_tables[sql_tables['TABLE_NAME'].str.contains(substring)]

    def concat_values(row)
    return f"{row['TABLE_SCHEMA']}.{row['TABLE_NAME']}"

    sql_tables['TABLE_NAME'] = sql_tables.apply(concat_values, axis=1)

    sql_tables = sql_tables.reset_index(drop=True)

    sql_tables = sql_tables[['TABLE_NAME']]

    print(tabulate(sql_tables, headers='keys', tablefmt='pretty'))


def list_columns(table, conn):

    schema, table = table.split('.')

    sql_types = pd.read_sql(f'exec sp_columns \'' +
                            table+'\', \''+schema+'\'', conn)

    sql_types = sql_types[['COLUMN_NAME', 'TYPE_NAME']]

    print(tabulate(sql_types, headers='keys', tablefmt='pretty'))


def create_script(view, conn):
    sql_create = pd.read_sql('EXEC sp_helptext \''+view+'\'', conn)
    print(tabulate(sql_create, headers='keys', tablefmt='pretty'))


def primary_keys(table, conn):
    schema, table = table.split('.')

    sql_primary_keys = pd.read_sql(
        f'exec sp_primarykeys \''+table+'\', \''+schema+'\'', conn)

    sql_primary_keys = sql_primary_keys[['COLUMN_NAME']]

    print(tabulate(sql_primary_keys, headers='keys', tablefmt='pretty'))


def upload_sql(df, table, conn):

    table_array = table.split('.')
    schema, table = table_array[0], table_array[1]

    try:
        df.to_sql(table, conn, if_exists='append', index=False)
        print(f'Uploaded {df} to {schema}.{table}')

    except:
        print(f'Failed to upload {df} to {schema}.{table}')
        print(traceback.format_exc())

    return

################ MANIPULATION FUNCTIONS ################


def lookup(lookup_value, lookup_array, return_array):
    match_value = return_array.loc[lookup_array == lookup_value]
    if match_value.empty:
        return None
    else:
        return match_value.tolist()[0]


def find_substring(df, substring):
    for i in range(len(df)):
        for j in range(len(df.iloc[i])):
            if substring in df.iloc[i][j]:
                return i, j


def normalize_string(string):
    if not isinstance(string, str):
        string = str(string)

    string = unicodedata.normalize('NFKD', string).encode('ascii', 'ignore')
    string = re.sub(r'[^\w\s]', '', string).strip().lower()
    string = ''.join(c for c in string if unicodedata.category(c) != 'Mn')
    return string


def only_letters(string):
    if not isinstance(string, str):
        string = str(string)

    string = s.lower()
    return re.sub('[^a-zA-Z]', '', string)

################ WRITE FUNCTIONS ################

# function that receives a dataframe as a parameter and returns its name as a string


def get_df_name(df):
    caller_globals = inspect.currentframe().f_back.f_globals

    for name, obj in caller_globals.items():
        if isinstance(obj, pd.DataFrame) and obj is df:
            return name
    return None


def export_to_excel(dataframes, filename):
    # Create a workbook
    workbook = openpyxl.Workbook()

    for df in dataframes:

        # get the name of the dataframe
        sheet_name = get_df_name(df)

        # Create a worksheet
        worksheet = workbook.create_sheet(sheet_name)

        # Write the data from the dataframe to the worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Set the column widths
        for column in df:
            column_length = max(df[column].astype(
                str).map(len).max(), len(column)) + 10
            col_idx = df.columns.get_loc(column)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                col_idx+1)].width = column_length

        # Freeze the header row
        worksheet.freeze_panes = worksheet['A2']

        # Enable filters for the worksheet
        num_rows, num_cols = df.shape
        cell_range = f'A1:{openpyxl.utils.get_column_letter(num_cols)}{num_rows}'
        worksheet.auto_filter.ref = cell_range

    del workbook['Sheet']

    # Save the workbook
    workbook.save(filename)

# function that receives a dataframe as parameter and prints its columns as well as the value in the first row using tabulate


def print_df(df):
    column_names = df.columns
    first_row = df.iloc[0].values
    data = {'column_name': column_names, 'first_row': first_row}
    print(tabulate(pd.DataFrame(data), headers='keys', tablefmt='pretty'))


################ OTHER FUNCTIONS ################


def get_latest_file(directory):
    files = os.listdir(directory)
    latest_file = max(files, key=os.path.getctime)
    return os.path.join(directory, latest_file)

################ USAGE ################

# ---------------- SQL ----------------#
# search_table('test', conn)
# list_columns('test', conn)
# create_script('test', conn)
# df = pd.read_sql('SELECT * FROM test', conn)
# upload_sql(df, 'test', conn)
# engine.execute('DELETE FROM table WHERE id = 1')

# ---------------- MANIPULATION ----------------#
# df = df.merge(df2[['id', 'name']], how='left', left_on='id', right_on='id')
# df = df.drop_duplicates(subset='id', keep='last', inplace=True)
# df1['value'] = df1['id'].apply(lookup, args = (df2['id'], df2['value']))
# print(tabulate(df1, headers='keys', tablefmt='pretty'))
# substring_location = find_substring(df, 'test')
# substring_row = substring_location[0]
# substring_col = substring_location[1]
# normalized_string = normalize_string(string)
# df['normalized_string'] = df['value'].apply(normalize_string)
# letters_only_string = letters_only(string)
# df['letters_only_string'] = df['value'].apply(letters_only)

# ---------------- WRITE ----------------#
# df = pd.read_csv('/path/to/file.csv', encoding='ansi', sep=';', lineterminator='\\r', encoding_errors='ignore', usecols=['id', 'name', 'value'])
# export_to_excel([df], '/path/to/file.xlsx')
# print_df(df)

# ---------------- OTHER ----------------#
# file_path = get_latest_file('/path/to/directory')
